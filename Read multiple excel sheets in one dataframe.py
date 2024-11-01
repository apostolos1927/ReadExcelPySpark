# Databricks notebook source
# MAGIC %pip install openpyxl

# COMMAND ----------

from openpyxl import load_workbook

excel_path = "/dbfs/FileStore/demo.xlsx"
 
# Load the workbook
#read only mode to be able to read huge excel files
workbook = load_workbook(excel_path, read_only=True)

# Function to read data from a single sheet
def read_sheet(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        #By setting values_only=True, you exclude cell formatting, formulas, and hyperlinks, which simplifies data handling when you only need the data content.
        if not any(row):
            continue  # skip rows that start with None
        data.append(row)
    return data


all_data_frames = []
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    data = read_sheet(sheet)

    # Assuming first row is the header
    headers = data[0]
    data_rows = data[1:]
    df = spark.createDataFrame(data_rows, schema=headers)
    all_data_frames.append(df)


final_df = all_data_frames[0]
for df in all_data_frames[1:]:
    final_df = final_df.unionByName(df)

final_df.display()
workbook.close()

# COMMAND ----------

# MAGIC %md
# MAGIC openpyxl does not support xls format, only xlsx.
